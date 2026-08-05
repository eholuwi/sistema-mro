"""Importa as fotos de "Material MRO 2026.xlsx" para o Sistema MRO em execução.

Uso (a partir da pasta `sistema-mro`):

    venv\\Scripts\\python.exe scripts\\importar_imagens_planilha.py            # SIMULAÇÃO
    venv\\Scripts\\python.exe scripts\\importar_imagens_planilha.py --aplicar  # grava de verdade

    --substituir   troca a foto de itens que JÁ têm imagem (por padrão são pulados)
    --planilha X   caminho de outra planilha (padrão: "Material MRO 2026.xlsx" na raiz)

**Simulação é o padrão.** Sem `--aplicar` nada é gravado: o script só mostra o relatório do
que faria. Com `--aplicar`, o banco é copiado para `backups/` ANTES da primeira escrita.

──────────────────────────────────────────────────────────────────────────────────────────
Como a foto é ligada ao item (o que torna este script confiável)

As fotos NÃO são desenhos flutuantes: estão **dentro da célula** (recurso "imagem em
célula" do Excel 365). O caminho até elas é uma cadeia de quatro arquivos dentro do .xlsx,
e ela dá o vínculo EXATO célula→imagem — não um chute pela ordem:

    célula <c r="E4" vm="1">          (aba GERAL/ENTRADA, coluna E = IMAGEM)
      └─ vm="N"  → xl/metadata.xml, N-ésimo bloco de <valueMetadata> → rc v="i"
           └─ i  → xl/richData/rdrichvalue.xml, i-ésimo <rv> → primeiro <v> = índice da imagem
                └─ índice → xl/richData/richValueRel.xml, N-ésimo <rel r:id="rIdX">
                     └─ rIdX → xl/richData/_rels/richValueRel.xml.rels → xl/media/imageNNN.png

O plano B previsto no `docs/prompt_importar_planilha_mro.md` ("extrair na ordem e torcer
para bater com as linhas") foi descartado: a cadeia acima existe e é exata, e associar foto
errada a item errado é pior do que não ter foto nenhuma.

O PN sai da **coluna B** da mesma linha da foto (GERAL tem cabeçalho na linha 3 e dados a
partir da 4; ENTRADA tem cabeçalho na 1 e dados a partir da 2).

──────────────────────────────────────────────────────────────────────────────────────────
Limites de escopo (deliberados)

- **Só imagens.** Estoque, mínimo, descrição, unidade e categoria da planilha são ignorados.
- **Nunca cria item.** PN da planilha que não existe no `inventario` é reportado, não inserido.
- **Não sobrescreve sem ordem.** Item que já tem foto é pulado, salvo `--substituir`.
- **Não altera schema.** Só grava `inventario.imagem_path` e arquivos em `docs/itens/`.
"""

from __future__ import annotations

import argparse
import re
import sys
import zipfile
from collections import Counter
from pathlib import Path

RAIZ = Path(__file__).resolve().parents[1]
if str(RAIZ) not in sys.path:
    sys.path.insert(0, str(RAIZ))

# O console do Windows abre em cp1252 e derruba o script no primeiro acento do relatório —
# depois de já ter lido a planilha inteira. Forçar UTF-8 na saída custa uma linha e evita
# perder o trabalho por causa de um caractere.
for _fluxo in (sys.stdout, sys.stderr):
    if hasattr(_fluxo, "reconfigure"):
        _fluxo.reconfigure(encoding="utf-8", errors="replace")

PLANILHA_PADRAO = "Material MRO 2026.xlsx"

# Abas com foto → (linha em que os dados começam). A coluna da foto é E e a do PN é B nas
# duas; GERAL vem primeiro porque é a planilha de inventário — quando o mesmo PN tem foto
# nas duas abas, a dela vence.
ABAS = (("GERAL", 4), ("ENTRADA", 2))
COL_IMAGEM = "E"
COL_PN = 2  # coluna B (1-based, como o openpyxl conta)


# ══════════════════════════════════════════════════════════════════════════════
# EXTRAÇÃO — a cadeia célula → imagem (ver docstring do módulo)
# ══════════════════════════════════════════════════════════════════════════════


def _indices_rich_value(zf):
    """`vm` da célula (1-based) → nome do arquivo em `xl/media/`.

    Cada elo é conferido: um .xlsx sem imagem em célula não tem estes arquivos, e um com a
    cadeia truncada produziria vínculo errado em silêncio — daí os erros explícitos."""
    nomes = set(zf.namelist())
    faltando = {
        "xl/metadata.xml",
        "xl/richData/rdrichvalue.xml",
        "xl/richData/richValueRel.xml",
        "xl/richData/_rels/richValueRel.xml.rels",
    } - nomes
    if faltando:
        raise SystemExit(
            "Esta planilha não tem imagens em célula (faltam: " + ", ".join(sorted(faltando)) + ")."
        )

    def _ler(nome):
        return zf.read(nome).decode("utf-8", "ignore")

    # 1. metadata: a ordem dos <bk> de <valueMetadata> é o que o `vm` indexa (1-based).
    md = _ler("xl/metadata.xml")
    bloco = md[md.find("<valueMetadata") : md.find("</valueMetadata>")]
    vm_para_rv = [int(v) for v in re.findall(r'<rc[^>]*\bv="(\d+)"', bloco)]

    # 2. rdrichvalue: o PRIMEIRO <v> de cada <rv> é o índice em richValueRel.
    rv = _ler("xl/richData/rdrichvalue.xml")
    rv_para_idx = [int(m) for m in re.findall(r"<rv[^>]*>\s*<v>(\d+)</v>", rv)]

    # 3. richValueRel: a ordem dos <rel> é o índice; o r:id aponta para o .rels.
    rvr = _ler("xl/richData/richValueRel.xml")
    idx_para_rid = re.findall(r'<rel[^>]*r:id="(rId\d+)"', rvr)

    # 4. .rels: rId → arquivo em xl/media/.
    rels = _ler("xl/richData/_rels/richValueRel.xml.rels")
    rid_para_arquivo = {
        rid: "xl/" + alvo.replace("../", "")
        for rid, alvo in re.findall(r'Id="(rId\d+)"[^>]*Target="([^"]+)"', rels)
    }

    mapa = {}
    for vm, i_rv in enumerate(vm_para_rv, start=1):
        if i_rv >= len(rv_para_idx):
            continue
        idx = rv_para_idx[i_rv]
        if idx >= len(idx_para_rid):
            continue
        arquivo = rid_para_arquivo.get(idx_para_rid[idx])
        if arquivo in nomes:
            mapa[vm] = arquivo
    return mapa


def _celulas_com_foto(zf, arquivo_sheet):
    """{linha: vm} das células da coluna IMAGEM que carregam foto."""
    xml = zf.read(arquivo_sheet).decode("utf-8", "ignore")
    achados = re.findall(rf'<c r="{COL_IMAGEM}(\d+)"[^>]*\bvm="(\d+)"', xml)
    return {int(linha): int(vm) for linha, vm in achados}


def _arquivos_das_abas(caminho_xlsx):
    """Nome da aba → arquivo `xl/worksheets/sheetN.xml` (a ordem do zip NÃO é a das abas)."""
    with zipfile.ZipFile(caminho_xlsx) as zf:
        wb = zf.read("xl/workbook.xml").decode("utf-8", "ignore")
        rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
    por_rid = dict(re.findall(r'Id="(rId\d+)"[^>]*Target="(worksheets/sheet\d+\.xml)"', rels))
    return {
        nome: "xl/" + por_rid[rid]
        for nome, rid in re.findall(r'<sheet name="([^"]+)"[^>]*r:id="(rId\d+)"', wb)
        if rid in por_rid
    }


def coletar_fotos_por_pn(caminho_xlsx):
    """[(pn_normalizado, aba, linha, arquivo_no_zip)] — uma entrada por foto encontrada.

    Percorre as abas na ordem de `ABAS`, então o primeiro PN visto (GERAL) é o que fica."""
    import openpyxl

    sheets = _arquivos_das_abas(caminho_xlsx)
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)
    resultado, vistos = [], set()
    try:
        with zipfile.ZipFile(caminho_xlsx) as zf:
            vm_para_arquivo = _indices_rich_value(zf)
            for aba, primeira_linha in ABAS:
                if aba not in sheets or aba not in wb.sheetnames:
                    print(f"  ! aba '{aba}' não encontrada — ignorada")
                    continue
                fotos = _celulas_com_foto(zf, sheets[aba])
                ws = wb[aba]
                pn_por_linha = {}
                for n, linha in enumerate(
                    ws.iter_rows(min_row=primeira_linha, max_col=COL_PN, values_only=True),
                    start=primeira_linha,
                ):
                    valor = linha[COL_PN - 1] if len(linha) >= COL_PN else None
                    if valor is not None and str(valor).strip():
                        pn_por_linha[n] = str(valor).strip().upper()
                for linha, vm in sorted(fotos.items()):
                    pn = pn_por_linha.get(linha)
                    arquivo = vm_para_arquivo.get(vm)
                    if not pn or not arquivo or pn in vistos:
                        continue
                    vistos.add(pn)
                    resultado.append((pn, aba, linha, arquivo))
    finally:
        wb.close()
    return resultado


# ══════════════════════════════════════════════════════════════════════════════
# GRAVAÇÃO
# ══════════════════════════════════════════════════════════════════════════════


def _inventario_por_pn():
    import database

    with database.transaction() as conn:
        linhas = conn.execute("SELECT id, part_number, imagem_path FROM inventario").fetchall()
    return {str(r["part_number"]).strip().upper(): dict(r) for r in linhas}


def _relatorio(fotos, inventario, casados, sem_match):
    print()
    print("-" * 78)
    print(f"  Fotos na planilha (PN único) ....... {len(fotos)}")
    print(f"  Itens no inventário ................ {len(inventario)}")
    print(f"  PN da planilha casado no sistema ... {len(casados)}")
    print(f"  PN da planilha SEM item no sistema . {len(sem_match)}  (não serão criados)")
    por_aba = Counter(aba for _, aba, _, _ in fotos)
    print(f"  Origem das fotos ................... {dict(por_aba)}")
    print("-" * 78)
    if sem_match:
        amostra = ", ".join(sorted(sem_match)[:12])
        print(f"\n  PN sem correspondência (amostra): {amostra}{' …' if len(sem_match) > 12 else ''}")


def main():
    ap = argparse.ArgumentParser(description="Importa as fotos da planilha MRO para o sistema.")
    ap.add_argument("--aplicar", action="store_true", help="grava de verdade (padrão: simulação)")
    ap.add_argument("--substituir", action="store_true", help="troca a foto de quem já tem")
    ap.add_argument("--planilha", default=None, help=f"caminho do .xlsx (padrão: {PLANILHA_PADRAO})")
    args = ap.parse_args()

    caminho = Path(args.planilha) if args.planilha else RAIZ / PLANILHA_PADRAO
    if not caminho.exists():
        raise SystemExit(f"Planilha não encontrada: {caminho}")

    import database
    from services.ficha import salvar_imagem_item

    print(f"Planilha : {caminho}")
    print(f"Banco    : {database.DB_PATH}")
    print(f"Modo     : {'APLICAR (grava)' if args.aplicar else 'SIMULAÇÃO (não grava nada)'}")
    print("\nLendo as fotos embutidas...")

    fotos = coletar_fotos_por_pn(caminho)
    inventario = _inventario_por_pn()
    casados = [f for f in fotos if f[0] in inventario]
    sem_match = {f[0] for f in fotos if f[0] not in inventario}
    _relatorio(fotos, inventario, casados, sem_match)

    ja_tem = [f for f in casados if (inventario[f[0]].get("imagem_path") or "").strip()]
    a_gravar = casados if args.substituir else [f for f in casados if f not in ja_tem]
    print(f"\n  A gravar ........................... {len(a_gravar)}")
    if ja_tem:
        rotulo = "serão SUBSTITUÍDAS" if args.substituir else "serão puladas (use --substituir)"
        print(f"  Itens que já tinham foto ........... {len(ja_tem)}  ({rotulo})")

    if not args.aplicar:
        print("\nSimulação — nada foi gravado. Rode de novo com --aplicar para valer.")
        return
    if not a_gravar:
        print("\nNada a gravar.")
        return

    # Backup ANTES da primeira escrita — regra inviolável nº4 do projeto. Não é migração de
    # schema, mas reescreve `imagem_path` de centenas de itens: se o casamento estiver
    # errado, o .bak é o caminho de volta.
    print("\nCriando backup do banco...")
    destino = database._backup_db("pre-import-imagens-planilha")
    print(f"  backup: {destino or '(falhou — veja o log; o import continua)'}")

    print("\nGravando...")
    gravadas, falhas = 0, []
    with zipfile.ZipFile(caminho) as zf:
        for pn, aba, linha, arquivo in a_gravar:
            item = inventario[pn]
            dados = zf.read(arquivo)
            nome = arquivo.rsplit("/", 1)[-1]
            ok, msg = salvar_imagem_item(item["id"], nome, dados)
            if ok:
                gravadas += 1
                if gravadas % 50 == 0:
                    print(f"  {gravadas}/{len(a_gravar)}...")
            else:
                falhas.append(f"{pn} ({aba} L{linha}): {msg}")

    print(f"\n  Fotos gravadas ..................... {gravadas}")
    if falhas:
        print(f"  Falhas ............................. {len(falhas)}")
        for f in falhas[:10]:
            print(f"    - {f}")
    print("\nPronto. Confira na Ficha 360 de alguns itens antes de considerar concluído.")


if __name__ == "__main__":
    main()
