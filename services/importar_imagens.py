"""Importa as fotos embutidas em "Material MRO 2026.xlsx" para o cadastro (v6.6.0).

Este módulo nasceu de `scripts/importar_imagens_planilha.py` (v2.6.0), que continua
existindo como CLI fino por cima daqui. A extração NÃO foi reescrita — foi movida, para
que a tela de Configurações e o script contem a mesma história.

──────────────────────────────────────────────────────────────────────────────────────
Como a foto é ligada ao item (o que torna este import confiável)

As fotos NÃO são desenhos flutuantes: estão **dentro da célula** (recurso "imagem em
célula" do Excel 365). O caminho até elas é uma cadeia de quatro arquivos dentro do
.xlsx, e ela dá o vínculo EXATO célula→imagem — não um chute pela ordem:

    célula <c r="E4" vm="1">          (aba GERAL/ENTRADA, coluna E = IMAGEM)
      └─ vm="N"  → xl/metadata.xml, N-ésimo bloco de <valueMetadata> → rc v="i"
           └─ i  → xl/richData/rdrichvalue.xml, i-ésimo <rv> → primeiro <v> = índice
                └─ índice → xl/richData/richValueRel.xml, N-ésimo <rel r:id="rIdX">
                     └─ rIdX → xl/richData/_rels/richValueRel.xml.rels → xl/media/…

O plano B ("extrair na ordem e torcer para bater com as linhas") foi descartado: a
cadeia acima existe e é exata, e associar foto errada a item errado é pior do que não
ter foto nenhuma.

O PN sai da **coluna B** da mesma linha da foto (GERAL tem cabeçalho na linha 3 e dados
a partir da 4; ENTRADA tem cabeçalho na 1 e dados a partir da 2).

──────────────────────────────────────────────────────────────────────────────────────
Limites de escopo (deliberados)

- **Só imagens.** Estoque, mínimo, descrição, unidade e categoria da planilha são
  ignorados — quem atualiza isso é `importar_inventario_neidson`.
- **Nunca cria item.** PN da planilha que não existe no `inventario` é reportado.
- **Não sobrescreve sem ordem.** Item que já tem foto é pulado, salvo `substituir=True`.
- **Não altera schema.** Só grava `inventario.imagem_path` e arquivos em `docs/itens/`.
"""

from __future__ import annotations

import re
import zipfile
from collections import Counter
from pathlib import Path

# Abas com foto → (linha em que os dados começam). A coluna da foto é E e a do PN é B nas
# duas; GERAL vem primeiro porque é a planilha de inventário — quando o mesmo PN tem foto
# nas duas abas, a dela vence.
ABAS = (("GERAL", 4), ("ENTRADA", 2))
COL_IMAGEM = "E"
COL_PN = 2  # coluna B (1-based, como o openpyxl conta)


class SemImagensEmCelula(ValueError):
    """A planilha não usa "imagem em célula" — não há o que importar.

    Erro explícito em vez de devolver zero fotos: um .xlsx com as fotos como desenhos
    flutuantes produziria "0 fotos encontradas" e mandaria o usuário procurar defeito no
    cadastro, quando o problema é o formato do arquivo.
    """


# ══════════════════════════════════════════════════════════════════════════════
# EXTRAÇÃO — a cadeia célula → imagem (ver docstring do módulo)
# ══════════════════════════════════════════════════════════════════════════════


def _indices_rich_value(zf):
    """`vm` da célula (1-based) → nome do arquivo em `xl/media/`.

    Cada elo é conferido: um .xlsx sem imagem em célula não tem estes arquivos, e um com
    a cadeia truncada produziria vínculo errado em silêncio — daí os erros explícitos."""
    nomes = set(zf.namelist())
    faltando = {
        "xl/metadata.xml",
        "xl/richData/rdrichvalue.xml",
        "xl/richData/richValueRel.xml",
        "xl/richData/_rels/richValueRel.xml.rels",
    } - nomes
    if faltando:
        raise SemImagensEmCelula(
            "Esta planilha não tem imagens em célula (faltam: " + ", ".join(sorted(faltando)) + ")."
        )

    def _ler(nome):
        return zf.read(nome).decode("utf-8", "ignore")

    # 1. metadata: a ordem dos <bk> de <valueMetadata> é o que o `vm` indexa (1-based).
    md = _ler("xl/metadata.xml")
    bloco = md[md.find("<valueMetadata") : md.find("</valueMetadata>")]
    vm_para_rv = [int(v) for v in re.findall(r'<rc[^>]*\bv="(\d+)"', bloco)]

    # 2. rdrichvalue: o PRIMEIRO <v> de cada <rv> é o índice em richValueRel.
    rv = _ler("xl/richData/rdrichvalue.xml")
    rv_para_idx = [int(m) for m in re.findall(r"<rv[^>]*>\s*<v>(\d+)</v>", rv)]

    # 3. richValueRel: a ordem dos <rel> é o índice; o r:id aponta para o .rels.
    rvr = _ler("xl/richData/richValueRel.xml")
    idx_para_rid = re.findall(r'<rel[^>]*r:id="(rId\d+)"', rvr)

    # 4. .rels: rId → arquivo em xl/media/ (mesmo leitor tolerante a ordem/atalho do
    #    passo anterior — ver `_relacionamentos`).
    rid_para_arquivo = _relacionamentos(_ler("xl/richData/_rels/richValueRel.xml.rels"))

    mapa = {}
    for vm, i_rv in enumerate(vm_para_rv, start=1):
        if i_rv >= len(rv_para_idx):
            continue
        idx = rv_para_idx[i_rv]
        if idx >= len(idx_para_rid):
            continue
        arquivo = rid_para_arquivo.get(idx_para_rid[idx])
        if arquivo in nomes:
            mapa[vm] = arquivo
    return mapa


def _celulas_com_foto(zf, arquivo_sheet):
    """{linha: vm} das células da coluna IMAGEM que carregam foto."""
    xml = zf.read(arquivo_sheet).decode("utf-8", "ignore")
    achados = re.findall(rf'<c r="{COL_IMAGEM}(\d+)"[^>]*\bvm="(\d+)"', xml)
    return {int(linha): int(vm) for linha, vm in achados}


def _atributo(tag: str, nome: str) -> str | None:
    m = re.search(rf'\b{re.escape(nome)}="([^"]*)"', tag)
    return m.group(1) if m else None


def _relacionamentos(xml: str) -> dict[str, str]:
    """`{rId: caminho dentro do zip}` de um `.rels`, SEM depender da ordem dos atributos.

    Ordem de atributo não tem significado em XML, e os produtores discordam: o Excel
    escreve `Id="rId1" Target="worksheets/sheet1.xml"`, o openpyxl escreve
    `Target="/xl/worksheets/sheet1.xml" Id="rId1"`. Uma regex que exigia `Id` primeiro
    devolvia dicionário VAZIO no segundo caso — e o sintoma era "0 fotos encontradas",
    que manda procurar defeito no cadastro em vez de no leitor. Também normaliza os três
    formatos de alvo vistos na prática: `../media/x.png`, `worksheets/s1.xml` e
    `/xl/worksheets/s1.xml`.
    """
    mapa = {}
    for tag in re.findall(r"<Relationship\b[^>]*>", xml):
        rid, alvo = _atributo(tag, "Id"), _atributo(tag, "Target")
        if not rid or not alvo:
            continue
        limpo = alvo.replace("../", "").lstrip("/")
        mapa[rid] = limpo if limpo.startswith("xl/") else "xl/" + limpo
    return mapa


def _arquivos_das_abas(caminho_xlsx):
    """Nome da aba → arquivo `xl/worksheets/sheetN.xml` (a ordem do zip NÃO é a das abas)."""
    with zipfile.ZipFile(caminho_xlsx) as zf:
        wb = zf.read("xl/workbook.xml").decode("utf-8", "ignore")
        rels = zf.read("xl/_rels/workbook.xml.rels").decode("utf-8", "ignore")
    por_rid = {
        rid: alvo
        for rid, alvo in _relacionamentos(rels).items()
        if re.search(r"worksheets/sheet\d+\.xml$", alvo)
    }
    abas = {}
    for tag in re.findall(r"<sheet\b[^>]*>", wb):
        nome, rid = _atributo(tag, "name"), _atributo(tag, "r:id")
        if nome and rid in por_rid:
            abas[nome] = por_rid[rid]
    return abas


def coletar_fotos_por_pn(caminho_xlsx):
    """[(pn_normalizado, aba, linha, arquivo_no_zip)] — uma entrada por foto encontrada.

    Percorre as abas na ordem de `ABAS`, então o primeiro PN visto (GERAL) é o que fica."""
    import openpyxl

    sheets = _arquivos_das_abas(caminho_xlsx)
    wb = openpyxl.load_workbook(caminho_xlsx, read_only=True, data_only=True)
    resultado, vistos = [], set()
    try:
        with zipfile.ZipFile(caminho_xlsx) as zf:
            vm_para_arquivo = _indices_rich_value(zf)
            for aba, primeira_linha in ABAS:
                if aba not in sheets or aba not in wb.sheetnames:
                    continue
                fotos = _celulas_com_foto(zf, sheets[aba])
                ws = wb[aba]
                pn_por_linha = {}
                for n, linha in enumerate(
                    ws.iter_rows(min_row=primeira_linha, max_col=COL_PN, values_only=True),
                    start=primeira_linha,
                ):
                    valor = linha[COL_PN - 1] if len(linha) >= COL_PN else None
                    if valor is not None and str(valor).strip():
                        pn_por_linha[n] = str(valor).strip().upper()
                for linha, vm in sorted(fotos.items()):
                    pn = pn_por_linha.get(linha)
                    arquivo = vm_para_arquivo.get(vm)
                    if not pn or not arquivo or pn in vistos:
                        continue
                    vistos.add(pn)
                    resultado.append((pn, aba, linha, arquivo))
    finally:
        wb.close()
    return resultado


def caminho_padrao_planilha() -> Path:
    """Palpite razoável para o campo de caminho da tela: `<pasta do banco>\\docs\\`.

    Em dev acerta em cheio (`sistema-mro\\docs\\Material MRO 2026.xlsx`); em produção
    aponta para `C:\\MRO\\dados\\docs\\`, ao lado de `docs\\itens\\`, que é onde as fotos
    dos itens já moram. É só um valor inicial — o campo é editável.
    """
    import database

    return Path(database.DB_PATH).resolve().parent / "docs" / "Material MRO 2026.xlsx"


def inventario_por_pn():
    """{PN normalizado: {id, part_number, imagem_path}} — o lado do sistema do casamento.

    Lê `inventario` inteiro de propósito, sem filtrar por `ativo`: casar foto é operação
    de cadastro, e item desativado continua tendo ficha."""
    import database

    with database.transaction() as conn:
        linhas = conn.execute("SELECT id, part_number, imagem_path FROM inventario").fetchall()
    return {str(r["part_number"]).strip().upper(): dict(r) for r in linhas}


# ══════════════════════════════════════════════════════════════════════════════
# ORQUESTRADOR — o contrato que a tela e o CLI consomem
# ══════════════════════════════════════════════════════════════════════════════


def importar_imagens_planilha(
    caminho,
    substituir: bool = False,
    dry_run: bool = True,
    fotos=None,
    progresso=None,
) -> tuple[bool, dict]:
    """Casa as fotos da planilha com o cadastro por Part Number. `(ok, stats)`.

    Mesmo contrato de `importar_inventario_neidson` — `dry_run=True` é o PADRÃO, e nesse
    modo nada é gravado: a tela mostra os números e só então oferece o Aplicar.

    `fotos` permite reaproveitar a coleta da pré-visualização: abrir um .xlsx de 118 MB
    com o openpyxl é o passo caro, e sem isso o usuário pagaria por ele duas vezes
    (uma na simulação, outra ao aplicar).

    `progresso` é um callable `(feitos, total)` — a gravação de ~400 arquivos leva tempo
    suficiente para uma tela muda parecer travada.
    """
    caminho = Path(caminho)
    stats: dict = {
        "fotos_na_planilha": 0,
        "itens_no_sistema": 0,
        "casados": 0,
        "sem_item_no_sistema": 0,
        "ja_tinham_foto": 0,
        "fotos_perdidas": 0,
        "a_gravar": 0,
        "gravadas": 0,
        "falhas": [],
        "pns_nao_encontrados": [],
        "por_aba": {},
        "dry_run": dry_run,
    }

    if not caminho.is_file():
        return False, {**stats, "erro": f"Arquivo não encontrado: {caminho}"}

    try:
        if fotos is None:
            fotos = coletar_fotos_por_pn(caminho)
    except SemImagensEmCelula as e:
        return False, {**stats, "erro": str(e)}
    except (zipfile.BadZipFile, KeyError) as e:
        return False, {**stats, "erro": f"Não consegui ler a planilha ({type(e).__name__}: {e})."}

    from services.ficha import imagem_existente

    inventario = inventario_por_pn()
    casados = [f for f in fotos if f[0] in inventario]
    sem_match = sorted({f[0] for f in fotos if f[0] not in inventario})

    # "Já tem foto" significa **o arquivo existe**, não "a coluna está preenchida".
    # `imagem_path` aponta para `docs/itens/`, que fica FORA do SQLite e viaja por
    # OneDrive: no banco de dev, 24 itens têm o caminho gravado e nenhum arquivo no disco.
    # Pela coluna, esses 24 seriam pulados para sempre e continuariam sem foto — falha
    # silenciosa. Pelo arquivo, o import os REPARA, que é o comportamento desejado
    # (regravar por cima de um caminho pendurado não destrói nada).
    ja_tem = [f for f in casados if imagem_existente(inventario[f[0]])]
    orfaos = [f for f in casados if f not in ja_tem and (inventario[f[0]].get("imagem_path") or "").strip()]
    a_gravar = casados if substituir else [f for f in casados if f not in ja_tem]

    stats.update(
        {
            "fotos_na_planilha": len(fotos),
            "itens_no_sistema": len(inventario),
            "casados": len(casados),
            "sem_item_no_sistema": len(sem_match),
            "ja_tinham_foto": len(ja_tem),
            "fotos_perdidas": len(orfaos),
            "a_gravar": len(a_gravar),
            "pns_nao_encontrados": sem_match,
            "por_aba": dict(Counter(aba for _, aba, _, _ in fotos)),
        }
    )

    if dry_run or not a_gravar:
        return True, stats

    # Backup ANTES da primeira escrita — regra inviolável nº4. Não é migração de schema,
    # mas reescreve `imagem_path` de centenas de itens: se o casamento estiver errado, o
    # .bak é o caminho de volta.
    import database
    from services.ficha import salvar_imagem_item

    stats["backup"] = str(database._backup_db("pre-import-imagens") or "")

    gravadas, falhas = 0, []
    total = len(a_gravar)
    with zipfile.ZipFile(caminho) as zf:
        for n, (pn, aba, linha, arquivo) in enumerate(a_gravar, start=1):
            item = inventario[pn]
            ok, msg = salvar_imagem_item(item["id"], arquivo.rsplit("/", 1)[-1], zf.read(arquivo))
            if ok:
                gravadas += 1
            else:
                falhas.append(f"{pn} ({aba} L{linha}): {msg}")
            if progresso is not None:
                progresso(n, total)

    stats["gravadas"] = gravadas
    stats["falhas"] = falhas
    return True, stats
