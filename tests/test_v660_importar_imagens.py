"""v6.6.0 — Import das fotos da planilha do MRO (extraido do CLI para servico).

O teste que justifica o arquivo e `test_foto_vai_para_o_pn_da_propria_linha`: a planilha
monta um .xlsx real com a cadeia rich-value do Excel 365 e liga as fotos ao CONTRARIO da
ordem das linhas. Se algum dia alguem trocar a cadeia por "extrai na ordem e torce para
bater", este teste falha — que e exatamente o erro que nao pode acontecer em silencio
(foto errada no item errado e pior do que item sem foto).
"""

from __future__ import annotations

import shutil
import sys
import zipfile
from pathlib import Path

import pytest

PROJ = Path(__file__).resolve().parents[1]
if str(PROJ) not in sys.path:
    sys.path.insert(0, str(PROJ))

from services.importar_imagens import (  # noqa: E402
    SemImagensEmCelula,
    coletar_fotos_por_pn,
    importar_imagens_planilha,
)

PNG1 = b"\x89PNG\r\n\x1a\n" + b"PRIMEIRA" * 4
PNG2 = b"\x89PNG\r\n\x1a\n" + b"SEGUNDA_" * 4


# ══════════════════════════════════════════════════════════════════════════════
# Planilha sintetica com "imagem em celula" (Excel 365)
# ══════════════════════════════════════════════════════════════════════════════

_METADATA = """<?xml version="1.0" encoding="UTF-8"?>
<metadata xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main">
<valueMetadata count="{n}">{bks}</valueMetadata></metadata>"""

_RDRICHVALUE = """<?xml version="1.0" encoding="UTF-8"?>
<rvData xmlns="http://schemas.microsoft.com/office/spreadsheetml/2017/richdata" count="{n}">
{rvs}</rvData>"""

_RICHVALUEREL = """<?xml version="1.0" encoding="UTF-8"?>
<richValueRels xmlns="http://schemas.microsoft.com/office/spreadsheetml/2022/richvaluerel"
 xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships">{rels}</richValueRels>"""

_RELS = """<?xml version="1.0" encoding="UTF-8"?>
<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">{rs}</Relationships>"""


def _planilha(destino: Path, pn_por_linha: dict[int, str], imagens_por_linha: dict[int, bytes]):
    """Monta um .xlsx com fotos EM CELULA na coluna E, ligadas pela cadeia rich-value.

    A ligacao e montada de tras para frente de proposito (a 1a celula aponta para a
    ULTIMA imagem): so a cadeia leva a foto certa ao PN certo — a ordem, nao.
    """
    import openpyxl

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "GERAL"
    ws["B3"] = "PART NUMBER"
    ws["E3"] = "IMAGEM"
    for linha, pn in pn_por_linha.items():
        ws.cell(row=linha, column=2, value=pn)
    for linha in imagens_por_linha:
        ws.cell(row=linha, column=5, value="#VALOR!")  # placeholder da imagem em celula
    bruto = destino.parent / "bruto.xlsx"
    wb.save(bruto)

    linhas = sorted(imagens_por_linha)
    n = len(linhas)
    # vm da celula (1-based) -> indice do <rv>, INVERTIDO: vm=1 aponta para o ultimo rv.
    vm_para_rv = list(range(n - 1, -1, -1))

    partes = {
        "xl/metadata.xml": _METADATA.format(
            n=n, bks="".join(f'<bk><rc t="1" v="{i}"/></bk>' for i in vm_para_rv)
        ),
        "xl/richData/rdrichvalue.xml": _RDRICHVALUE.format(
            n=n, rvs="".join(f'<rv s="0"><v>{i}</v></rv>' for i in range(n))
        ),
        "xl/richData/richValueRel.xml": _RICHVALUEREL.format(
            rels="".join(f'<rel r:id="rId{i + 1}"/>' for i in range(n))
        ),
        "xl/richData/_rels/richValueRel.xml.rels": _RELS.format(
            rs="".join(
                f'<Relationship Id="rId{i + 1}" Type="http://img" Target="../media/image{i + 1}.png"/>'
                for i in range(n)
            )
        ),
    }
    for i, linha in enumerate(linhas, start=1):
        partes[f"xl/media/image{i}.png"] = imagens_por_linha[linha]

    with zipfile.ZipFile(bruto) as origem, zipfile.ZipFile(destino, "w", zipfile.ZIP_DEFLATED) as saida:
        for item in origem.infolist():
            dados = origem.read(item.filename)
            if item.filename.startswith("xl/worksheets/sheet"):
                texto = dados.decode("utf-8")
                for vm, linha in enumerate(linhas, start=1):
                    texto = texto.replace(f'<c r="E{linha}"', f'<c r="E{linha}" vm="{vm}"', 1)
                dados = texto.encode("utf-8")
            saida.writestr(item.filename, dados)
        for nome, conteudo in partes.items():
            saida.writestr(nome, conteudo if isinstance(conteudo, bytes) else conteudo.encode("utf-8"))
    return destino


@pytest.fixture
def planilha(tmp_path):
    """Duas fotos, nas linhas 4 e 5, para os PNs PN-A e PN-B."""
    return _planilha(
        tmp_path / "MRO.xlsx",
        {4: "PN-A", 5: "PN-B"},
        {4: PNG1, 5: PNG2},
    )


# ══════════════════════════════════════════════════════════════════════════════
# Extracao — a cadeia celula -> imagem
# ══════════════════════════════════════════════════════════════════════════════


def test_coleta_encontra_as_fotos(planilha):
    fotos = coletar_fotos_por_pn(planilha)
    assert {f[0] for f in fotos} == {"PN-A", "PN-B"}
    assert {f[1] for f in fotos} == {"GERAL"}
    assert {f[2] for f in fotos} == {4, 5}


def test_foto_vai_para_o_pn_da_propria_linha(planilha):
    """A cadeia liga E4 a ULTIMA imagem — a ordem levaria a foto errada ao item errado."""
    por_pn = {pn: arquivo for pn, _, _, arquivo in coletar_fotos_por_pn(planilha)}
    assert por_pn["PN-A"] == "xl/media/image2.png"
    assert por_pn["PN-B"] == "xl/media/image1.png"

    with zipfile.ZipFile(planilha) as zf:
        assert zf.read(por_pn["PN-A"]) == PNG2
        assert zf.read(por_pn["PN-B"]) == PNG1


def test_planilha_sem_imagem_em_celula(tmp_path):
    """Erro explicito: "0 fotos" mandaria procurar defeito no cadastro, nao no arquivo."""
    import openpyxl

    wb = openpyxl.Workbook()
    wb.active.title = "GERAL"
    wb.active["B4"] = "PN-A"
    caminho = tmp_path / "sem_fotos.xlsx"
    wb.save(caminho)

    with pytest.raises(SemImagensEmCelula):
        coletar_fotos_por_pn(caminho)


def test_pn_normalizado_por_trim_e_maiuscula(tmp_path):
    caminho = _planilha(tmp_path / "MRO.xlsx", {4: "  pn-a  "}, {4: PNG1})
    assert coletar_fotos_por_pn(caminho)[0][0] == "PN-A"


# ══════════════════════════════════════════════════════════════════════════════
# Orquestrador — casamento com o cadastro
# ══════════════════════════════════════════════════════════════════════════════


def test_contrato_e_dry_run_nao_grava(db, make_item, planilha):
    make_item(part_number="PN-A")
    ok, stats = importar_imagens_planilha(planilha, dry_run=True)

    assert ok
    assert stats["dry_run"] is True
    assert stats["fotos_na_planilha"] == 2
    assert stats["casados"] == 1
    assert stats["sem_item_no_sistema"] == 1
    assert stats["a_gravar"] == 1
    assert stats["gravadas"] == 0

    conn = db.get_connection()
    linha = conn.execute("SELECT imagem_path FROM inventario WHERE part_number='PN-A'").fetchone()
    conn.close()
    assert linha["imagem_path"] is None, "dry_run nao pode escrever nada"


def test_aplicar_grava_arquivo_e_imagem_path(db, make_item, planilha):
    item_id = make_item(part_number="PN-A")
    ok, stats = importar_imagens_planilha(planilha, dry_run=False)

    assert ok and stats["gravadas"] == 1 and stats["falhas"] == []

    conn = db.get_connection()
    linha = conn.execute("SELECT imagem_path FROM inventario WHERE id=?", (item_id,)).fetchone()
    conn.close()
    assert linha["imagem_path"] == f"docs/itens/item_{item_id}.png"

    disco = Path(db.DB_PATH).parent / "docs" / "itens" / f"item_{item_id}.png"
    assert disco.read_bytes() == PNG2, "PN-A tem que receber a foto da PROPRIA linha"


def test_pn_sem_cadastro_e_reportado_e_nunca_criado(db, make_item, planilha):
    make_item(part_number="PN-A")
    ok, stats = importar_imagens_planilha(planilha, dry_run=False)

    assert ok
    assert stats["pns_nao_encontrados"] == ["PN-B"]

    conn = db.get_connection()
    n = conn.execute("SELECT COUNT(*) AS n FROM inventario").fetchone()["n"]
    conn.close()
    assert n == 1, "o import nunca cria item"


def test_item_com_foto_no_disco_e_pulado(db, make_item, planilha):
    item_id = make_item(part_number="PN-A")
    importar_imagens_planilha(planilha, dry_run=False)
    disco = Path(db.DB_PATH).parent / "docs" / "itens" / f"item_{item_id}.png"
    disco.write_bytes(b"\x89PNG\r\n\x1a\nFOTO_MANUAL")

    ok, stats = importar_imagens_planilha(planilha, dry_run=False)

    assert ok and stats["ja_tinham_foto"] == 1 and stats["a_gravar"] == 0
    assert disco.read_bytes() == b"\x89PNG\r\n\x1a\nFOTO_MANUAL", "foto manual foi sobrescrita"


def test_substituir_troca_a_foto_existente(db, make_item, planilha):
    item_id = make_item(part_number="PN-A")
    importar_imagens_planilha(planilha, dry_run=False)
    disco = Path(db.DB_PATH).parent / "docs" / "itens" / f"item_{item_id}.png"
    disco.write_bytes(b"\x89PNG\r\n\x1a\nFOTO_MANUAL")

    ok, stats = importar_imagens_planilha(planilha, substituir=True, dry_run=False)

    assert ok and stats["gravadas"] == 1
    assert disco.read_bytes() == PNG2


def test_foto_cadastrada_mas_sumida_do_disco_e_reparada(db, make_item, planilha):
    """`imagem_path` preenchido com o arquivo ausente NAO conta como "ja tem foto".

    Medido no banco real em 12/08/2026: 24 itens com o caminho gravado e ZERO arquivos
    no disco (`docs/itens/` fica fora do SQLite e viaja por OneDrive). Pelo criterio da
    coluna, esses itens seriam pulados para sempre e continuariam sem foto — falha
    silenciosa. O criterio certo e a existencia do ARQUIVO.
    """
    item_id = make_item(part_number="PN-A")
    importar_imagens_planilha(planilha, dry_run=False)
    disco = Path(db.DB_PATH).parent / "docs" / "itens" / f"item_{item_id}.png"
    disco.unlink()  # o arquivo some, o imagem_path fica

    ok, stats = importar_imagens_planilha(planilha, dry_run=False)

    assert ok
    assert stats["ja_tinham_foto"] == 0
    assert stats["fotos_perdidas"] == 1
    assert stats["gravadas"] == 1, "item com foto pendurada tem que ser reparado"
    assert disco.read_bytes() == PNG2


def test_backup_do_banco_antes_de_gravar(db, make_item, planilha):
    """Regra inviolavel no4: reescrever `imagem_path` de centenas de itens exige o .bak."""
    make_item(part_number="PN-A")
    ok, stats = importar_imagens_planilha(planilha, dry_run=False)

    assert ok and stats["backup"]
    assert list((Path(db.DB_PATH).parent / "backups").glob("*pre-import-imagens*"))


def test_arquivo_inexistente_devolve_erro(db, tmp_path):
    ok, stats = importar_imagens_planilha(tmp_path / "nao_existe.xlsx")
    assert not ok and "não encontrado" in stats["erro"]


def test_fotos_prontas_nao_reabrem_a_planilha(db, make_item, planilha, monkeypatch):
    """A coleta e o passo caro (openpyxl num .xlsx de 118 MB) — a tela a reusa entre a
    pre-visualizacao e o Aplicar, senao o usuario paga por ela duas vezes."""
    import services.importar_imagens as M

    make_item(part_number="PN-A")
    fotos = coletar_fotos_por_pn(planilha)

    def _explode(*_a, **_k):
        raise AssertionError("coletar_fotos_por_pn foi chamada de novo")

    monkeypatch.setattr(M, "coletar_fotos_por_pn", _explode)

    ok, stats = M.importar_imagens_planilha(planilha, dry_run=False, fotos=fotos)
    assert ok and stats["casados"] == 1 and stats["gravadas"] == 1


def test_progresso_e_chamado_por_foto(db, make_item, planilha):
    make_item(part_number="PN-A")
    passos = []
    importar_imagens_planilha(planilha, dry_run=False, progresso=lambda f, t: passos.append((f, t)))
    assert passos == [(1, 1)]


# ══════════════════════════════════════════════════════════════════════════════
# O CLI nao pode ter uma segunda copia da logica
# ══════════════════════════════════════════════════════════════════════════════


def test_cli_delega_ao_servico():
    fonte = (PROJ / "scripts" / "importar_imagens_planilha.py").read_text(encoding="utf-8")
    assert "from services.importar_imagens import" in fonte
    for termo in ("_indices_rich_value", "richValueRel", "def coletar_fotos_por_pn"):
        assert termo not in fonte, f"a cadeia de extracao voltou para o CLI (`{termo}`)"


def test_caminho_padrao_aponta_para_ao_lado_do_banco(db):
    from services.importar_imagens import caminho_padrao_planilha

    esperado = Path(db.DB_PATH).resolve().parent / "docs" / "Material MRO 2026.xlsx"
    assert caminho_padrao_planilha() == esperado


@pytest.fixture(autouse=True)
def _limpa_itens_dir(db):
    """Cada teste comeca sem `docs/itens/` (o servico grava ao lado do banco isolado)."""
    yield
    alvo = Path(db.DB_PATH).parent / "docs"
    if alvo.exists():
        shutil.rmtree(alvo, ignore_errors=True)
